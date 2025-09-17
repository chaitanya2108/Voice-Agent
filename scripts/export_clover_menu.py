import os
import sys
from openpyxl import Workbook

# Ensure project root is on sys.path for imports when running as a script
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from restaurant_data import restaurant_context


def map_variants(category: str, item_name: str):
    # Default: no variants
    has_variants = "No"
    attribute = ""
    option_names = ""

    # Pizzas: sizes
    if category == "pizza":
        has_variants = "Yes"
        attribute = "size"
        option_names = "Small;Medium;Large"
    # Pasta: Regular/Large
    elif category == "pasta":
        has_variants = "Yes"
        attribute = "size"
        option_names = "Regular;Large"
    # Appetizers: Regular/Party via custom
    elif category == "appetizers":
        has_variants = "Yes"
        attribute = "custom"
        option_names = "Regular;Party"
    # Beverages specifics
    elif category == "beverages":
        if item_name == "Gelato":
            has_variants = "Yes"
            attribute = "flavour"
            option_names = "Vanilla;Chocolate;Strawberry"
        elif item_name in {"Italian Soda", "Cappuccino"}:
            has_variants = "Yes"
            attribute = "size"
            option_names = "Small;Regular;Large"

    return has_variants, attribute, option_names


def export_xlsx(output_path: str):
    # Create workbook with EXACT sheet order required by Clover bulk import
    wb = Workbook()

    # Remove default sheet and re-create to ensure order
    default = wb.active
    wb.remove(default)

    # 1) Items sheet
    items_ws = wb.create_sheet("Items")
    items_headers = [
        "Name",            # Item name
        "Price",           # Price in dollars
        "Category",        # Category name
        "Description",     # Optional
        "Allergens",       # Optional
        "Has Variants",    # Yes/No
        "Attribute",       # size|flavour|color|custom
        "Option Names"     # Semicolon separated
    ]
    items_ws.append(items_headers)

    # Populate items
    for category, items in restaurant_context.menu.items():
        for item in items:
            allergens = "|".join(item.get("allergens", [])) if item.get("allergens") else ""
            has_variants, attribute, option_names = map_variants(category, item["name"])
            items_ws.append([
                item["name"],
                float(item["price"]),
                category,
                item.get("description", ""),
                allergens,
                has_variants,
                attribute,
                option_names,
            ])

    # 2) Modifier Groups sheet (leave empty for now with headers)
    mg_ws = wb.create_sheet("Modifier Groups")
    mg_headers = [
        "Group Name",
        "Modifier Name",
        "Price Delta"
    ]
    mg_ws.append(mg_headers)

    # 3) Categories sheet
    cat_ws = wb.create_sheet("Categories")
    cat_headers = ["Category Name"]
    cat_ws.append(cat_headers)
    for category in restaurant_context.menu.keys():
        cat_ws.append([category])

    # 4) Tax Rates sheet
    tax_ws = wb.create_sheet("Tax Rates")
    tax_headers = ["Tax Name", "Rate Percent"]
    tax_ws.append(tax_headers)
    # Provide default 8.5% tax consistent with app
    tax_ws.append(["Default Tax", 8.5])

    wb.save(output_path)


if __name__ == "__main__":
    out_file = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "clover_menu_items.xlsx"))
    export_xlsx(out_file)
    print(f"Saved: {out_file}")
